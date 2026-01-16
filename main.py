print("Starting CVE Database...")

import sqlite3
import csv
import json
import os
import requests
import time

DB_NAME = "data/cvedb.sqlite3"
DATA_DIR = "data"


# Init DB
def init_db():
    conn = sqlite3.connect(DB_NAME)
    with open("schema.sql", "r", encoding="utf-8") as f:
        conn.executescript(f.read())
    conn.close()
    print("✅ Database initialized")


# Import CSV

def bulk_import_csv(file_path):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    inserted = 0

    with open(file_path, newline="", encoding="utf-8") as csvfile:
        reader = csv.DictReader(csvfile)
        reader.fieldnames = [h.strip() for h in reader.fieldnames]

        for row in reader:
            cve_id = row.get("cve_id", "").strip()
            if not cve_id:
                continue

            try:
                cursor.execute("""
                    INSERT INTO cves
                    (cve_id, description, published_date, severity, cvss_score, vendor, product)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    cve_id,
                    row.get("description", ""),
                    row.get("published_date", ""),
                    row.get("severity", ""),
                    None,
                    row.get("vendor", "UNKNOWN"),
                    row.get("affected_product", "UNKNOWN")
                ))
                inserted += 1
            except sqlite3.IntegrityError:
                pass

    conn.commit()
    conn.close()
    print(f"   → Imported {inserted} from {os.path.basename(file_path)}")


# Import NVD JSON

def bulk_import_nvd_json(file_path):
    import json
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    inserted = 0

    with open(file_path, encoding="utf-8") as f:
        data = json.load(f)

    for item in data.get("vulnerabilities", []):
        cve = item.get("cve", {})
        cve_id = cve.get("id", "")
        if not cve_id:
            continue

        # Description
        descs = cve.get("descriptions", [])
        description = descs[0]["value"] if descs else ""

        # Published date
        published = cve.get("published", "")[:10]

        # CVSS
        severity = "UNKNOWN"
        score = None
        metrics = cve.get("metrics", {})

        if "cvssMetricV31" in metrics:
            m = metrics["cvssMetricV31"][0]["cvssData"]
            severity = m["baseSeverity"]
            score = m["baseScore"]
        elif "cvssMetricV30" in metrics:
            m = metrics["cvssMetricV30"][0]["cvssData"]
            severity = m["baseSeverity"]
            score = m["baseScore"]
        elif "cvssMetricV2" in metrics:
            m = metrics["cvssMetricV2"][0]
            severity = m["baseSeverity"]
            score = m["cvssData"]["baseScore"]

        # Vendor + Product
        vendor = "UNKNOWN"
        product = "UNKNOWN"

        configs = cve.get("configurations", [])
        if configs:
            nodes = configs[0].get("nodes", [])
            if nodes:
                matches = nodes[0].get("cpeMatch", [])
                if matches:
                    cpe = matches[0]["criteria"]
                    parts = cpe.split(":")
                    if len(parts) > 4:
                        vendor = parts[3]
                        product = parts[4]

        try:
            cursor.execute("""
                INSERT INTO cves
                (cve_id, description, published_date, severity, cvss_score, vendor, product)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (cve_id, description, published, severity, score, vendor, product))
            inserted += 1
        except sqlite3.IntegrityError:
            pass

    conn.commit()
    conn.close()
    print(f"✅ Imported {inserted} CVEs from {file_path}")



# Auto import all file form data

def import_all_data():
    print("\n Scanning data folder...")

    for file in os.listdir(DATA_DIR):
        path = os.path.join(DATA_DIR, file)

        if file.endswith(".sqlite3"):
            continue

        elif file.lower().endswith(".csv"):
            print(f" CSV → {file}")
            bulk_import_csv(path)

        elif file.lower().endswith(".json"):
            print(f" JSON → {file}")
            bulk_import_nvd_json(path)

    print("\n✅ All files imported")

# live fetching latest CVEs from NVD

# live fetching latest CVEs from NVD





# Add CVEs
def add_cve():
    cve_id = input("CVE ID: ")
    desc = input("Description: ")
    date = input("Published date: ")
    sev = input("Severity: ")
    prod = input("Product: ")
    vendor = input("Vendor: ")

    conn = sqlite3.connect(DB_NAME)
    try:
        conn.execute(
            "INSERT INTO cves VALUES (?, ?, ?, ?, ?, ?)",
            (cve_id, desc, date, sev, prod, vendor)
        )
        conn.commit()
        print(" CVE added")
    except sqlite3.IntegrityError:
        print("⚠ CVE already exists")
    conn.close()


# Find CVE
def find_cve():
    cid = input("Enter CVE ID: ")
    conn = sqlite3.connect(DB_NAME)
    row = conn.execute("SELECT * FROM cves WHERE cve_id = ?", (cid,)).fetchone()
    conn.close()

    if row:
        print("\n--- CVE FOUND ---")
        for col in row:
            print(col)
    else:
        print(" Not Availabe")

# View all

def view_all():
    conn = sqlite3.connect(DB_NAME)
    rows = conn.execute("SELECT cve_id, severity FROM cves LIMIT 50").fetchall()
    conn.close()

    for r in rows:
        print(r)

def search_by_vendor():
    v = input("Vendor name: ").lower()
    conn = sqlite3.connect(DB_NAME)
    rows = conn.execute(
        "SELECT cve_id, severity, product FROM cves WHERE LOWER(vendor)=?",
        (v,)
    ).fetchall()
    conn.close()

    for r in rows:
        print(r)

def search_critical():
    conn = sqlite3.connect(DB_NAME)
    rows = conn.execute(
        "SELECT cve_id, cvss_score, vendor, product FROM cves WHERE severity='CRITICAL' ORDER BY cvss_score DESC"
    ).fetchall()
    conn.close()

    for r in rows:
        print(r)

def export_to_csv():
    import csv
    output_file = input("Enter output CSV file name (default: cve_report.csv): ").strip()
    if not output_file:
        output_file = "cve_report.csv"

    conn = sqlite3.connect(DB_NAME)
    rows = conn.execute("SELECT * FROM cves ORDER BY published_date DESC").fetchall()
    conn.close()

    if not rows:
        print(" No CVEs in database to export.")
        return

    headers = ["cve_id", "description", "published_date", "severity", "cvss_score", "vendor", "product"]

    with open(output_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)

    print(f" Exported {len(rows)} CVEs to {output_file}")

def export_to_xlsx():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    output_file = input("Enter XLSX file name (default: cve_report.xlsx): ").strip()
    if not output_file:
        output_file = "cve_report.xlsx"

    conn = sqlite3.connect(DB_NAME)
    rows = conn.execute("SELECT * FROM cves ORDER BY published_date DESC").fetchall()
    conn.close()

    if not rows:
        print(" No CVEs in database to export.")
        return

    headers = ["CVE ID", "Description", "Published Date", "Severity", "CVSS Score", "Vendor", "Product"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CVE Report"

    # Write headers with formatting
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Severity colors
    severity_colors = {
        "CRITICAL": "FF0000",  # Red
        "HIGH": "FF6600",      # Orange
        "MEDIUM": "FFFF00",    # Yellow
        "LOW": "00FF00",       # Green
        "UNKNOWN": "FFFFFF"    # White
    }

    # Write CVE rows
    for row_num, row in enumerate(rows, 2):
        for col_num, value in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num, value=value)
        sev = str(row[3]).upper() if row[3] else "UNKNOWN"
        fill_color = severity_colors.get(sev, "FFFFFF")
        for col_num in range(1, len(headers)+1):
            ws.cell(row=row_num, column=col_num).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Freeze top row
    ws.freeze_panes = "A2"

    wb.save(output_file)
    print(f" Exported {len(rows)} CVEs to {output_file}")


 
# Main Menu

while True:
    print("\n===== CVE DATABASE =====")
    print("1. Initialize DB")
    print("2. Add CVE")
    print("3. Find CVE")
    print("4. View CVEs")
    print("5. Import ALL files from data folder")
    print("6. Search by Vendor")
    print("7. Show Critical CVEs")
    print("8. Export to CSV")
    print("9. Export to XLSX")
    print("10. Exit")


    choice = input("Choose: ")

    if choice == "1":
        init_db()
    elif choice == "2":
        add_cve()
    elif choice == "3":
        find_cve()
    elif choice == "4":
        view_all()
    elif choice == "5":
        import_all_data()
    elif choice == "6":
        search_by_vendor()
    elif choice == "7":
        search_critical()
    elif choice == "8":
        export_to_csv()
    elif choice == "9":
        export_to_xlsx()
    elif choice == "10":
        print("Good Bye...")
        break